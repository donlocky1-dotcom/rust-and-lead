#!/usr/bin/env python3
# Baut die vollstaendige Asset-Liste fuer "Rust & Lead" als Excel-Arbeitsmappe.
# Quellen: godot/scripts/{CombatData,ProgressionManager,WorldManager,EquipManager}.gd,
#          docs/MASTER_GDD.md, docs/STORY_BIBLE.md, index.html (Prototyp).
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/workspace/rust-and-lead/docs/ASSET_LISTE.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="2F3E46")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
CAT_FILL = PatternFill("solid", fgColor="D8CBB4")
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="2F3E46")
NOTE_FONT = Font(name=FONT, size=10, italic=True, color="555555")
DONE_FILL = PatternFill("solid", fgColor="D5E8D4")   # bereits vorhanden
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (Kategorie, Name, Typ, Beschreibung/Referenz, Prio, Zielmass, Zielpfad, Registry-Name, Status)
# Typ: "3D-Modell" | "Textur" | "Icon 2D" | "VFX" | "Bild 2D"
# Prio: 1 = jetzt noetig (Spiel sichtbar unvollstaendig), 2 = bald, 3 = spaeter/Politur
A = []

def add(cat, name, typ, desc, prio, size, path, reg="", status="offen"):
    A.append([cat, name, typ, desc, prio, size, path, reg, status])

# ── SPIELER & BEGLEITER ────────────────────────────────────────────────────────
add("Spieler", "Spieler-Chassis (Einheit 13, nach dem Erwachen)", "3D-Modell",
    "Rauchender, einaeugiger Stahlschaedel, Zahnraeder, Schulterventile. Hauptfigur ab Kap. 4 (Story-Bibel Kap. 4.2).",
    1, "1,8 m", "assets/models/characters/player.glb", "player")
add("Spieler", "Spieler-Fleisch-Illusion (vor dem Erwachen)", "3D-Modell",
    "Gleiches Rig, menschliche Huelle. Wird beim Reveal abgeworfen. Kann Material-Variante des Chassis sein.",
    2, "1,8 m", "assets/models/characters/player_human.glb", "player_human")
add("Spieler", "Bolzen (Blechhund, Begleiter)", "3D-Modell",
    "Brass-Koerper, Schlappohr, Antennenschwanz, cyan Auge. Stoesst in Providence Cut dazu.",
    2, "0,7 m", "assets/models/characters/bolzen.glb", "companion_dog")
add("Spieler", "Bolzen — Wrack (zerstoert)", "3D-Modell",
    "Zusammengesunkenes Chassis, als Schrott-Pickup einsammelbar.",
    3, "0,5 m", "assets/models/characters/bolzen_wreck.glb", "companion_wreck")

# ── GEGNER (CombatData.ENEMY_TYPES) ───────────────────────────────────────────
add("Gegner", "Grenzgaenger (outlaw)", "3D-Modell",
    "BIOLOGISCH — rund + fleischrot (Kampf-Lesbarkeit GDD §8.4). Standard-Nahkaempfer.",
    1, "1,6 m", "assets/models/enemies/outlaw.glb", "enemy_outlaw")
add("Gegner", "Oelfresser-Ratte (fauna)", "3D-Modell",
    "BIOLOGISCH, Schwarm-Gegner, schnell und klein.",
    1, "0,6 m", "assets/models/enemies/fauna.glb", "enemy_fauna")
add("Gegner", "Revolverheld (revolver)", "3D-Modell",
    "BIOLOGISCH, Fernkaempfer — haelt Abstand und feuert.",
    1, "1,6 m", "assets/models/enemies/revolver.glb", "enemy_revolver")
add("Gegner", "Konzern-Konstrukt (konstrukt)", "3D-Modell",
    "MECHANISCH — eckig + stahlblau. Dampf aus Nackenventilen, Gesicht wie zugeschweisste Ofenklappe.",
    1, "1,9 m", "assets/models/enemies/konstrukt.glb", "enemy_konstrukt")
add("Gegner", "Kessel-Klaeffer (klaeffer)", "3D-Modell",
    "MECHANISCH, Schwarm — mechanischer Hund, sehr schnell.",
    1, "0,8 m", "assets/models/enemies/klaeffer.glb", "enemy_klaeffer")
add("Gegner", "Schwerer Ernter / Goliath", "3D-Modell",
    "MECHANISCH, Boss. Frontal panzerimmun (Saeure noetig). Kolben-Telegraph-Angriffe.",
    2, "4,0 m", "assets/models/enemies/goliath.glb", "enemy_goliath")
add("Gegner", "Minen-Titan (Superboss)", "3D-Modell",
    "Dungeon-Titan der Schrott-Minen, Ebene 3.",
    3, "6,0 m", "assets/models/enemies/minentitan.glb", "enemy_minentitan")
add("Gegner", "Schrott-Golem (Superboss)", "3D-Modell",
    "Superboss; garantierter Golem-Faust-Drop.",
    3, "6,0 m", "assets/models/enemies/schrottgolem.glb", "enemy_schrottgolem")
add("Gegner", "Champion-Aura / Unique-Kennzeichnung", "VFX",
    "Visuelle Aufwertung fuer benannte Champions (30 % in Kritter-Hallen) — Aura + groessere Silhouette.",
    3, "—", "assets/vfx/champion_aura", "")

# ── BOSSE / STORY-GEGNER ──────────────────────────────────────────────────────
add("Boss", "Kolben-Jack (Zugmeister, Kap. 4)", "3D-Modell",
    "2 m Vorarbeiter in Konzerngrau, rechter Arm = dampfgetriebene Ramme.",
    2, "2,2 m", "assets/models/enemies/kolben_jack.glb", "boss_kolbenjack")
add("Boss", "Direktor Cornelius Vane — Phase 1 (Wachs)", "3D-Modell",
    "Vaeterliches Wachsgesicht, makelloser Anzug. Endgegner Kap. 12.",
    3, "1,9 m", "assets/models/enemies/vane_p1.glb", "boss_vane1")
add("Boss", "Direktor Vane — Phase 2 (Direktoren-Chassis)", "3D-Modell",
    "Wachs abgeschmolzen, freigelegtes Stahl-Chassis darunter.",
    3, "2,2 m", "assets/models/enemies/vane_p2.glb", "boss_vane2")
add("Boss", "Das Eiserne Herz (Phase 3, Maschinenhalle)", "3D-Modell",
    "Hausgrosser Kolbenmotor, wird selbst zum Gegner. Herzschlag-Motiv.",
    3, "25 m", "assets/models/enemies/eisernes_herz_boss.glb", "boss_herz")

# ── NPCs (Story-Bibel) ────────────────────────────────────────────────────────
for nm, desc, prio in [
    ("Mamma „Rusty“ Mabel", "Wirtin des Gatling-Saloons, ~50, Unterarme wie eine Schmiedin. Zentrale Bezugsperson.", 1),
    ("Silas „Kupferauge“ Finch", "Schmied, mechanische Kupferlinse statt Auge, Teilprothesen.", 1),
    ("Doktor „Doc“ Aris", "Feldarzt, hager, Nickelbrille (putzt sie, wenn er luegt).", 1),
    ("Pip", "Waisenkind, 9, Zahnluecke, Schrottsammlerin. Haelt den Helden fuer ihren Schutzengel.", 2),
    ("Gideon Cross", "Rebellenfuehrer, Narbe vom Kinn uebers Ohr. Kanonische Fraktion.", 2),
    ("Aufseher Quentin", "Eiserne Gilde, Uniform ohne Staubkorn, glatt und kuehl.", 3),
    ("„Slick“ Sterling", "Schmuggler-Hehler, bessere Weste als der ganze Canyon.", 3),
    ("Elias Roan", "Providence Cut, Familien-Bogen (roter Faden).", 2),
    ("Tessa „Spulen-Tess“ Marlow", "Signalmeisterin der Rebellen, redet schnell.", 3),
    ("Agata Dorn", "Haelt den Garten aus Rost an der Smog-Linie.", 3),
    ("Elias Thorn (Uhrmacher)", "Nebenstory B1 „Das Uhrwerk-Herz“, Trauerflor am Zylinder.", 3),
    ("Oliver (Kinder-Automat)", "Thorns Automat, Nebenstory B1.", 3),
    ("Corah Bell (Witwe)", "Nebenstory B2 „Die Witwe und der Automat“.", 3),
    ("Vesper Kane (Kopfgeldjaegerin)", "Nebenstory B3, Duster aus Goliath-Leder.", 3),
    ("Bruder Habakuk", "Nebenstory B4, aufgenaehte Haut ueber Stahl (Grusel-Ton).", 3),
    ("„Laterne“ Lomax (blinder Projektionist)", "Nebenstory B5, Augen weiss wie Milchglas.", 3),
    ("Dynamit-Dolores", "Nebenstory B7, Lunte als Haarband. Legendaere Waffe.", 3),
    ("Konzern-Patrouillenfuehrer", "Graue Uniform, liest von Messing-Lochkarten ab.", 2),
    ("Stadtbewohner (3 Varianten)", "Generische Rustwater-Bevoelkerung fuer Belebung.", 2),
]:
    add("NPC", nm, "3D-Modell", desc, prio, "1,7-1,8 m",
        "assets/models/npcs/", "")

# ── WAFFEN (CombatData.WEAPONS) ───────────────────────────────────────────────
for nm, dt, desc in [
    ("Blei-Karabiner", "KINETISCH", "Startwaffe, physische Munition."),
    ("Leydener Volt-Karabiner", "GALVANISCH", "Energiekristalle; stark gegen Maschinen (2,5x)."),
    ("Saeure-Spruecher", "ALCHEMISCH", "Zersetzt Panzerung (Korrosion)."),
    ("Dampf-Brenner", "THERMISCH", "Setzt Ziele in Brand (DOT)."),
]:
    add("Waffe", nm, "3D-Modell", f"{dt}. In der Hand des Spielers sichtbar + als Loot-Objekt.",
        1, "0,9-1,2 m", "assets/models/weapons/", "")

# ── AUSRUESTUNG (ProgressionManager.GEAR_SLOTS) ───────────────────────────────
for nm, stat, desc in [
    ("Helm", "Leben", "Sichtbar am Chassis-Kopf."),
    ("Ruestung", "Ruestung", "Torso-Panzerplatten."),
    ("Waffen-Gear", "Schaden", "Aufsatz/Lauf-Variante."),
    ("Gadget", "Feuerrate", "Guertel-/Schulteranbau."),
    ("Stiefel", "Tempo", "Hydraulik-Laufbeine."),
    ("Panzerplatte", "Ruestung", "Einsteckbare Platte (Plate-Slots)."),
]:
    add("Ausruestung", f"{nm} — 4 Seltenheitsstufen", "3D-Modell",
        f"Haupt-Stat {stat}. Je Stufe (Gewoehnlich/Selten/Episch/Legendaer) eine erkennbare Variante.",
        2, "slot-abhaengig", "assets/models/gear/", "")

for nm in ["Schaden-Chip", "Nachlade-Servo", "Laufwerk-Modul", "Vital-Kern", "Panzer-Kern"]:
    add("Ausruestung", f"Tech-Modul: {nm}", "3D-Modell",
        "Einsteckbares Modul fuer Platten-Slots (Diablo-Immortal-Gem-Prinzip).",
        3, "0,15 m", "assets/models/gear/tech/", "")

# ── LEGENDARIES (ProgressionManager.LEGENDARIES) ──────────────────────────────
for nm, slot, power in [
    ("Dolores' letzte Trommel", "Waffe", "Spezialschuss feuert 11 statt 7 Projektile"),
    ("Iron-Rail-Durchschlag", "Waffe", "Kugeln durchschlagen einen zusaetzlichen Gegner"),
    ("Galvanische Trommel", "Waffe", "Krits loesen Kurzschluss aus"),
    ("Ballistischer Rechenkern", "Waffe", "Krits prallen zu zweitem Ziel ab"),
    ("Golem-Faust", "Waffe", "+18 % Schaden (Boss-Drop Schrott-Golem)"),
    ("Titan-Kolben-Panzer", "Ruestung", "Boss-Flaechenschlaege halbiert"),
    ("Blutdampf-Kessel", "Ruestung", "Jeder Kill heilt 3 %"),
    ("Wachsherz-Kuerass", "Ruestung", "-15 % Schaden (Boss-Drop Vane)"),
    ("Sparventil-Uhr", "Gadget", "25 % kein Munitionsverbrauch"),
    ("Selbstschmier-Ventil", "Gadget", "Schnellere Regeneration"),
    ("Pluenderer-Sohlen", "Stiefel", "+25 % Gold, staerkerer Magnet"),
    ("Quecksilber-Sohlen", "Stiefel", "+12 % Tempo"),
    ("Kupferlinsen-Visier", "Helm", "+8 % Grund-Kritchance"),
    ("Kesselschaedel-Haube", "Helm", "+15 % max. Leben"),
]:
    add("Legendaer", nm, "3D-Modell",
        f"Slot {slot}. Kraft: {power}. Eigene, unverwechselbare Silhouette + Leucht-Akzent.",
        3, "slot-abhaengig", "assets/models/gear/legendary/", "")

# ── GEBAEUDE ──────────────────────────────────────────────────────────────────
for nm, desc, prio in [
    ("Gatling-Saloon", "Township-Kern. Ausgeschlachtete Gatling als Kronleuchter, Klavier.", 1),
    ("Eiserne Schmiede", "Silas' Werkstatt, Amboss, Esse, Funkenflug.", 1),
    ("Mondschein-Destille", "Township-Gebaeude (Wirtschafts-System).", 2),
    ("Alchemie-Labor / Raffinerie", "Township + Smog-Gate (Filter ab Stufe 3).", 2),
    ("Wohnhuette (3 Varianten)", "Bahnschwellen + Kesselblech, windschief.", 1),
    ("Wasserturm", "Rustwaters Wiedererkennungs-Silhouette, Teer-Schriftzug mit tropfendem W.", 1),
    ("Palisaden-Segment + Tor", "Stadtbefestigung, modular kachelbar.", 1),
    ("Bahnhof (Schnellreise)", "Nur an echten Hubs (GDD §1.4).", 2),
    ("Fort Freedom — Module", "Rebellenfestung aus Schienenstahl, Brustwehr.", 2),
    ("Sektor 01 — Module", "Konzern-Architektur, Kontrollhalle mit Manometern.", 3),
    ("Rogue's Landing — Module", "Schmugglerhoehle, Gaslicht.", 3),
    ("Eisernes Herz — Aussenbau", "Zentrale Landmarke, von ueberall sichtbar.", 2),
    ("Providence Cut — Ruinen-Module", "Verkohlte Balken wie Rippen, Ortsschild „214 Seelen“.", 2),
    ("Grabstein (3x, Familie)", "Liv, Sara, Tom — Familien-Bogen.", 2),
    ("Erinnerungspunkt / Memorial", "Interaktionspunkt Providence Cut.", 3),
]:
    add("Gebaeude", nm, "3D-Modell", desc, prio, "modular", "assets/models/buildings/", "")

# ── UMGEBUNG / FLORA ──────────────────────────────────────────────────────────
for nm, desc, prio, size, reg, status in [
    ("Kaktus (Wueste)", "Leit-Flora Sektor 1.", 1, "1-3 m", "flora_cactus", "offen"),
    ("Baum (Oase / Rostwald)", "Leit-Flora Gruene Senke + Rostwald.", 2, "4-7 m", "flora_tree", "offen"),
    ("Wasserflaeche (Oase)", "Gruene Senke.", 3, "—", "flora_water", "offen"),
    ("Salzkruste (Salzpfanne)", "Boden-Dekor Salzpfanne.", 3, "flach", "flora_salt", "offen"),
    ("Toter Baum (Smog-Oedland)", "Leit-Flora Sektor 3.", 2, "4-6 m", "flora_deadtree", "offen"),
    ("Fels klein", "Vorhanden (CC0 sand_rocks_small_01).", 1, "0,5 m", "rock_small", "vorhanden"),
    ("Felsbrocken", "Vorhanden (CC0 namaqualand_boulder_03).", 1, "1,5 m", "rock_boulder", "vorhanden"),
    ("Klippe", "Vorhanden (CC0 namaqualand_cliff_02).", 1, "7 m", "cliff", "vorhanden"),
    ("Kraterrand-Felswand (Modul)", "Aussengrenze der Welt, 350 m hoch, kachelbar.", 2, "modular", "crater_rim", "offen"),
    ("Rand-Tunnel-Tor", "Der eine Ausgang, konzern-bewacht (GDD §1.7.4).", 3, "60x80 m", "rim_tunnel", "offen"),
    ("Strassensegment + Kreuzung", "Handelsrouten zwischen Hubs, GridMap-tauglich.", 1, "9 m breit", "road_segment", "offen"),
    ("Bahngleis-Segment", "Iron-Rail-Trasse.", 2, "modular", "rail_segment", "offen"),
    ("Panzerzug: Lok + Waggon + Salonwagen", "Kap.-4-Bosskampf-Arena (Reveal-Szene).", 3, "gross", "train", "offen"),
]:
    add("Umgebung", nm, "3D-Modell", desc, prio, size, "assets/models/environment/", reg, status)

# ── PROPS ─────────────────────────────────────────────────────────────────────
for nm, desc, prio, reg, status in [
    ("Schatztruhe", "Vorhanden (CC0 treasure_chest).", 1, "chest", "vorhanden"),
    ("Munitionskiste", "Vorhanden (CC0 ammo_box).", 2, "ammo_box", "vorhanden"),
    ("Werkzeugwagen", "Vorhanden (CC0 tool_cart).", 3, "tool_cart", "vorhanden"),
    ("Metallregal", "Vorhanden (CC0 worn_metal_rack).", 3, "metal_rack", "vorhanden"),
    ("Industrie-Wandlampe", "Vorhanden (CC0 industrial_wall_lamp).", 3, "wall_lamp", "vorhanden"),
    ("Chemie-Set", "Vorhanden (CC0 chemistry_set).", 3, "chemistry_set", "vorhanden"),
    ("Stielgranate", "Vorhanden (CC0 stick_grenade).", 3, "grenade", "vorhanden"),
    ("Fass / Tonne", "Allgemeines Dekor.", 2, "barrel", "offen"),
    ("Holzkiste", "Allgemeines Dekor, stapelbar.", 2, "crate", "offen"),
    ("Kutschenwrack", "Pips Nebenstory + Wuesten-Dekor.", 3, "wagon_wreck", "offen"),
    ("Kinetoskop-Projektor", "Story-Schluesselobjekt (Familien-Botschaften, Faelschung).", 2, "projector", "offen"),
    ("Messingwalze (Erinnerungswalze)", "16 Stueck sammelbar; Walzen-Archiv im Finale.", 2, "memory_roll", "offen"),
    ("Walzen-Archiv-Regal", "Kathedrale aus Regalen, Finale Kap. 12.", 3, "archive_shelf", "offen"),
    ("Auftragsbrett (Kopfgeld)", "Saloon-Interaktion.", 2, "bounty_board", "offen"),
    ("Persoenliche Truhe (Stash)", "Lager in jeder Stadt; auch Bolzen-Anbindepunkt.", 2, "stash", "offen"),
]:
    add("Prop", nm, "3D-Modell", desc, prio, "0,3-2 m", "assets/models/props/", reg, status)

# ── ITEMS (Welt-Objekt + Icon) ────────────────────────────────────────────────
for nm, desc in [
    ("Schrott", "Grundmaterial, haeufigster Drop."),
    ("Zahnrad", "Mittleres Material."),
    ("Dampfkern", "Seltenes Material, Quest-Waehrung."),
    ("Heiltrank", "Verbrauchsgegenstand."),
    ("Munition (Blei)", "Karabiner-Munition, stapelbar bis 200."),
    ("Energiekristall", "Munition der Energiewaffen."),
    ("Goldmuenze", "Waehrung, Loot-Magnet-Ziel."),
    ("Tragesystem (Beute-Upgrade)", "Erweitert Inventar dauerhaft, nur als Beute."),
]:
    add("Item", nm, "3D-Modell", f"{desc} Weltobjekt zum Aufsammeln.", 2, "0,1-0,3 m",
        "assets/models/items/", "")
    add("Item", f"{nm} — Icon", "Icon 2D", f"{desc} Inventar-/HUD-Darstellung.", 1, "128x128 px",
        "assets/ui/icons/items/", "")

# ── UI / ICONS ────────────────────────────────────────────────────────────────
for nm, desc, prio, size in [
    ("Schadensart-Icons (4x)", "Kinetisch / Galvanisch / Alchemisch / Thermisch.", 1, "128x128 px"),
    ("Waffen-Icons (4x)", "Fuer Waffenumschalter im HUD.", 1, "128x128 px"),
    ("Slot-Icons (6x)", "Helm, Ruestung, Waffe, Gadget, Stiefel, Platte — leere Slots.", 1, "128x128 px"),
    ("Status-Icons (4x)", "Kurzschluss-Stun, Verbluten, Ueberhitzung, Korrosion.", 1, "64x64 px"),
    ("Klassen-Marker (2x)", "Eckig+stahlblau = Maschine, rund+fleischrot = organisch (GDD §8.4).", 1, "64x64 px"),
    ("Seltenheits-Rahmen (4x)", "Gewoehnlich/Selten/Episch/Legendaer — Inventar-Kacheln.", 1, "256x256 px"),
    ("Perk-Icons (12x)", "3 Zweige (Revolverheld/Kesseltreiber/Grenzgaenger) x 4 Perks.", 2, "128x128 px"),
    ("Faehigkeits-Buttons (5x)", "Spezialschuss, Ausweich-Dash, Heiltrank, Saeure-Granate, Elektro-Granate.", 1, "192x192 px"),
    ("Gilden-Wappen (3x)", "Rebellen, Eiserne Gilde, Schmuggler.", 2, "256x256 px"),
    ("Iron-Rail-Emblem", "Zahnrad, durchbohrt von einer Schiene. An jeder dritten Wand.", 1, "512x512 px"),
    ("Virtueller Joystick", "Basis + Daumen-Grafik (Mobile-First, GDD §1.5).", 1, "256x256 px"),
    ("HUD-Rahmen & Balken", "Leben, Kesseldruck, XP-Leiste, Boss-Leiste.", 1, "9-slice"),
    ("Minikarten-Symbole", "POI-Marker je Sektor, Spieler-Pfeil, Gegner-Punkt, Quest-Marker.", 2, "64x64 px"),
    ("Codex-Illustrationen", "Je Codex-Eintrag ein Bild (aktuell ~12 Eintraege).", 3, "1024x768 px"),
    ("Kapitel-Titelkarten (12x)", "Kapitel-Blenden der Kampagne.", 3, "1920x1080 px"),
    ("Logo / Titelbild", "Startbildschirm „Rust & Lead“.", 2, "1920x1080 px"),
    ("Ladebildschirm-Motive", "Mehrere Motive mit Story-Zitaten.", 3, "1920x1080 px"),
]:
    add("UI", nm, "Icon 2D" if "px" in size else "Bild 2D", desc, prio, size, "assets/ui/", "")

# ── TEXTUREN ──────────────────────────────────────────────────────────────────
for nm, desc, prio, reg, status in [
    ("Boden: Wueste / Sand", "Vorhanden (CC0 gravelly_sand), gekachelt ueber 5000 m.", 1, "ground_sand", "vorhanden"),
    ("Boden: Salzpfanne", "Weisse, rissige Kruste.", 2, "ground_salt", "offen"),
    ("Boden: Rostwald", "Rostrote Erde mit Metallspaenen.", 2, "ground_rostwald", "offen"),
    ("Boden: Kupfer-Hochland", "Kupfergruen verwittertes Gestein.", 2, "ground_kupfer", "offen"),
    ("Boden: Smog-Oedland", "Vergiftete, graugruene Erde.", 2, "ground_smog", "offen"),
    ("Boden: Oase / Gras", "Gruene Senke.", 3, "ground_oasis", "offen"),
    ("Material: Holz (verwittert)", "Gebaeude, Palisaden, Bahnschwellen.", 1, "mat_wood", "offen"),
    ("Material: Wellblech / Rost", "Huetten, Daecher.", 1, "mat_corrugated", "offen"),
    ("Material: Messing / Kupfer", "Maschinen, Walzen, Prothesen.", 1, "mat_brass", "offen"),
    ("Material: Stahl / Chassis", "Spieler, Konstrukte, Iron Rail.", 1, "mat_steel", "offen"),
    ("Material: Fels", "Kraterrand, Klippen.", 2, "mat_rock", "offen"),
    ("Skybox: Gruener Bronzehimmel", "Kanonische Himmelsfarbe (Story-Bibel). Tag/Nacht-faehig.", 1, "sky_bronze", "offen"),
]:
    add("Textur", nm, "Textur", desc, prio, "1K-2K PBR", "assets/textures/", reg, status)

# ── VFX ───────────────────────────────────────────────────────────────────────
for nm, desc, prio in [
    ("Muendungsfeuer (4x je Schadensart)", "Kinetisch/Galvanisch/Alchemisch/Thermisch, je eigene Farbe.", 1),
    ("Leuchtspur / Projektil (4x)", "Passend zur Schadensart (bereits farbcodiert im Code).", 1),
    ("Treffer-Impact (4x)", "Funken, Blitz, Aetzen, Flamme.", 1),
    ("Kurzschluss-Blitz (Stun)", "Galvanik-Statuseffekt, 4 s.", 1),
    ("Verbluten / Oel-Austritt", "DOT-Statuseffekt organisch/mechanisch.", 2),
    ("Ueberhitzung / Brennen", "Thermik-DOT.", 2),
    ("Korrosion (Saeure)", "Panzerung zersetzt sich sichtbar.", 2),
    ("Dampf-Ausstoss (Ventile)", "Das „Atemholen“ des Chassis — Kernmotiv.", 1),
    ("Explosion", "Granaten, Bosskampf, Sprengungen.", 2),
    ("Smog-Nebel (volumetrisch)", "Smog-Senke, toedlich ohne Filter.", 2),
    ("Loot-Strahl je Seltenheit", "Saeule ueber gefallenem Item, Farbe = Seltenheit.", 2),
    ("Level-Up / Perk-Effekt", "Aufstiegs-Feedback.", 2),
    ("Reveal-Sequenz (Spiegelsplitter)", "Die wichtigste Szene des Spiels, Kap. 4.", 3),
    ("Fog of War", "Volumetrischer Nebel, Sichtlinien-basiert (GDD §1.4).", 3),
    ("Tag/Nacht-Beleuchtung", "Godot-Lichtzyklus ueber der Overworld.", 3),
]:
    add("VFX", nm, "VFX", desc, prio, "—", "assets/vfx/", "")

# ── ANIMATIONEN ───────────────────────────────────────────────────────────────
for nm, desc, prio in [
    ("Spieler: Idle / Laufen / Schiessen / Dash / Treffer / Tod", "Kern-Animationsset des Chassis.", 1),
    ("Gegner: Idle / Laufen / Angriff / Tod (je Typ)", "Fuer alle 6 Gegner-Grundtypen.", 1),
    ("Bolzen: Laufen / Beissen / Schiessen / Ausfallen", "Begleiter-Animationen.", 3),
    ("NPC: Idle / Gestik / Sprechen", "Wiederverwendbares Set fuer alle NPCs.", 2),
    ("Boss-Telegraphs", "Angekuendigte Flaechenschlaege (Goliath, Vane, Jack).", 3),
]:
    add("Animation", nm, "Animation", desc, prio, "—", "assets/animations/", "")

# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Blatt 1: Uebersicht ───────────────────────────────────────────────────────
ov = wb.active
ov.title = "Uebersicht"
ov["A1"] = "RUST & LEAD — Asset-Liste"
ov["A1"].font = TITLE_FONT
ov["A2"] = "Alles, was grafisch ins Spiel soll: 3D-Modelle, Texturen, 2D-Icons, VFX und Animationen."
ov["A2"].font = NOTE_FONT
ov["A3"] = "Quellen: godot/scripts/*.gd (Kampf-, Item-, Welt-Daten), docs/MASTER_GDD.md, docs/STORY_BIABEL"
ov["A3"] = "Quellen: godot/scripts/*.gd (Kampf-, Item-, Welt-Daten), docs/MASTER_GDD.md, docs/STORY_BIBLE.md, index.html (Prototyp)."
ov["A3"].font = NOTE_FONT

ov["A5"] = "PRIORITAETEN"
ov["A5"].font = Font(name=FONT, bold=True, size=12)
for r, (p, txt) in enumerate([
    (1, "Jetzt noetig — ohne diese Assets wirkt das Spiel sichtbar unfertig (Platzhalter-Primitives)."),
    (2, "Bald — rundet Welt und Systeme ab, sobald Prioritaet 1 steht."),
    (3, "Spaeter / Politur — Story-Setpieces, Nebenstories, Feinschliff."),
], start=6):
    ov[f"A{r}"] = p
    ov[f"A{r}"].font = Font(name=FONT, bold=True)
    ov[f"B{r}"] = txt
    ov[f"B{r}"].font = Font(name=FONT)

ov["A10"] = "WORKFLOW (siehe godot/assets/README.md)"
ov["A10"].font = Font(name=FONT, bold=True, size=12)
for r, txt in enumerate([
    "In Blender entstehen einzelne, wiederverwendbare Module — nicht die fertige Welt.",
    "In Godot wird zusammengesetzt: GridMap zum „Malen“ von Strassen/Haeuserzeilen, Path3D fuer Kurven.",
    "Datei am Zielpfad ablegen — die AssetRegistry findet sie automatisch, sonst bleibt der Platzhalter.",
    "Ausrichtung: +Y oben, Blickrichtung -Z, Pivot am Boden (Fuesse auf Y=0). Groesse egal (Auto-Skalierung).",
], start=11):
    ov[f"A{r}"] = "•"
    ov[f"B{r}"] = txt
    ov[f"B{r}"].font = Font(name=FONT)

ov["A16"] = "ZUSAMMENFASSUNG"
ov["A16"].font = Font(name=FONT, bold=True, size=12)

# Kopfzeile der Auswertung
ov["A17"] = "Asset-Typ"; ov["B17"] = "Prio 1"; ov["C17"] = "Prio 2"; ov["D17"] = "Prio 3"
ov["E17"] = "Gesamt"; ov["F17"] = "davon vorhanden"; ov["G17"] = "offen"
for c in "ABCDEFG":
    ov[f"{c}17"].font = HDR_FONT
    ov[f"{c}17"].fill = HDR_FILL
    ov[f"{c}17"].border = BORDER

typen = ["3D-Modell", "Textur", "Icon 2D", "Bild 2D", "VFX", "Animation"]
LIST = "'Asset-Liste'"
LAST = len(A) + 1   # letzte Datenzeile der Liste (Kopfzeile + Eintraege)
for i, t in enumerate(typen):
    r = 18 + i
    ov[f"A{r}"] = t
    ov[f"B{r}"] = f'=COUNTIFS({LIST}!$C$2:$C${LAST},$A{r},{LIST}!$E$2:$E${LAST},1)'
    ov[f"C{r}"] = f'=COUNTIFS({LIST}!$C$2:$C${LAST},$A{r},{LIST}!$E$2:$E${LAST},2)'
    ov[f"D{r}"] = f'=COUNTIFS({LIST}!$C$2:$C${LAST},$A{r},{LIST}!$E$2:$E${LAST},3)'
    ov[f"E{r}"] = f'=COUNTIF({LIST}!$C$2:$C${LAST},$A{r})'
    ov[f"F{r}"] = f'=COUNTIFS({LIST}!$C$2:$C${LAST},$A{r},{LIST}!$I$2:$I${LAST},"vorhanden")'
    ov[f"G{r}"] = f'=E{r}-F{r}'
    for c in "ABCDEFG":
        ov[f"{c}{r}"].font = Font(name=FONT)
        ov[f"{c}{r}"].border = BORDER

rs = 18 + len(typen)
ov[f"A{rs}"] = "SUMME"
ov[f"A{rs}"].font = Font(name=FONT, bold=True)
for c in "BCDEFG":
    ov[f"{c}{rs}"] = f'=SUM({c}18:{c}{rs-1})'
    ov[f"{c}{rs}"].font = Font(name=FONT, bold=True)
for c in "ABCDEFG":
    ov[f"{c}{rs}"].fill = CAT_FILL
    ov[f"{c}{rs}"].border = BORDER

ov.column_dimensions["A"].width = 22
ov.column_dimensions["B"].width = 14
for c in "CDEFG":
    ov.column_dimensions[c].width = 16
ov.column_dimensions["B"].width = 95 if False else 14

# ── Blatt 2: Asset-Liste ──────────────────────────────────────────────────────
ws = wb.create_sheet("Asset-Liste")
headers = ["Kategorie", "Name", "Asset-Typ", "Beschreibung / Referenz", "Prio",
           "Zielmass", "Zielpfad in Godot", "Registry-Name", "Status"]
ws.append(headers)
for i, _ in enumerate(headers, start=1):
    c = ws.cell(row=1, column=i)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BORDER

for row in A:
    ws.append(row)

for r in range(2, ws.max_row + 1):
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=r, column=col)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=(col == 4))
        c.border = BORDER
    if ws.cell(row=r, column=9).value == "vorhanden":
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).fill = DONE_FILL

widths = [15, 42, 13, 68, 6, 14, 34, 20, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"

wb.save(OUT)
print("gespeichert:", OUT)
print("Zeilen:", len(A))
from collections import Counter
print("nach Typ:", dict(Counter(x[2] for x in A)))
print("nach Prio:", dict(Counter(x[4] for x in A)))
print("vorhanden:", sum(1 for x in A if x[8] == "vorhanden"))
